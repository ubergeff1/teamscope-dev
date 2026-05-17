"""
Projects router — full CRUD for projects, deliverables, phases, and workshops.

API prefix: /api/projects
Tags: ["projects"]

This router handles:
  - Project CRUD (list, create, get, update, delete)
  - Managing consultant assignments to projects (add/remove)
  - Deliverable CRUD within a project, including auto-date calculation
  - Reordering deliverables via drag-and-drop
  - Phase updates within deliverables
  - Workshop CRUD within a project
  - Syncing weekly allocations for all deliverables in a project
  - Auto-generating phase dates (forward from start or backward from end)
  - Exporting a full project to a multi-sheet Excel workbook
"""
from typing import Annotated
from datetime import date as dt_date, timedelta
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models.project import Project
from app.models.consultant import Consultant
from app.models.deliverable import Deliverable, Workshop, DeliverablePhase
from app.schemas.project import ProjectCreate, ProjectUpdate, ProjectOut
from app.schemas.deliverable import (
    DeliverableCreate, DeliverableUpdate, DeliverableOut,
    WorkshopCreate, WorkshopUpdate, WorkshopOut,
    PhaseUpdate, PhaseOut, DeliverableReorder,
)
from app.utils.auth import get_current_user
from app.utils.assignment_sync import sync_deliverable_assignments



def _add_business_days(start: dt_date, days: int) -> dt_date:
    """Add N business days to a start date (skipping weekends).

    Iterates forward from ``start``, counting only Monday-Friday days,
    until ``days`` business days have been added.

    Args:
        start: The date to start counting from.
        days: Number of business days to add. If <= 0, returns start unchanged.

    Returns:
        The resulting date after adding the specified number of business days.
    """
    if days <= 0:
        return start
    current = start
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Mon-Fri
            added += 1
    return current


def _sub_business_days(end: dt_date, days: int) -> dt_date:
    """Subtract N business days from an end date (skipping weekends).

    Iterates backward from ``end``, counting only Monday-Friday days,
    until ``days`` business days have been subtracted.

    Args:
        end: The date to start counting backward from.
        days: Number of business days to subtract. If <= 0, returns end unchanged.

    Returns:
        The resulting date after subtracting the specified number of business days.
    """
    if days <= 0:
        return end
    current = end
    removed = 0
    while removed < days:
        current -= timedelta(days=1)
        if current.weekday() < 5:
            removed += 1
    return current

# Create the FastAPI router with /projects prefix; all endpoints require authentication
router = APIRouter(prefix="/projects", tags=["projects"])

# Type aliases for dependency injection — used as type hints in endpoint signatures
# Auth: extracts the current authenticated username from the JWT token
Auth = Annotated[str, Depends(get_current_user)]
# DB: provides a SQLAlchemy database session, automatically closed after the request
DB = Annotated[Session, Depends(get_db)]

# Standard phase types created for each deliverable.
# "workshop" phase is only created for workshop-type deliverables.
_PHASE_TYPES = ["workshop", "draft", "qa", "delivery"]


def _get_project_or_404(db: Session, project_id: int) -> Project:
    """Fetch a project by primary key or raise HTTP 404.

    Args:
        db: The database session.
        project_id: The primary key of the project to retrieve.

    Returns:
        The Project ORM instance.

    Raises:
        HTTPException: 404 if the project does not exist.
    """
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    return p


# ── Projects ──────────────────────────────────────────────────────────────────

@router.get("", response_model=list[ProjectOut])
def list_projects(db: DB, _: Auth, status: str | None = None):
    """GET /projects — List all projects, optionally filtered by status.

    Query Parameters:
        status (str, optional): Filter projects by status value (e.g. "active", "complete").

    Returns:
        List of ProjectOut objects, ordered alphabetically by name.
    """
    q = db.query(Project)
    if status:
        q = q.filter(Project.status == status)
    return q.order_by(Project.name).all()


@router.post("", response_model=ProjectOut, status_code=201)
def create_project(body: ProjectCreate, db: DB, _: Auth):
    """POST /projects — Create a new project.

    Request Body:
        ProjectCreate schema with fields like name, client_name, status, dates, etc.

    Returns:
        The newly created ProjectOut object (HTTP 201).

    Business Logic:
        1. Deserialize the request body into a Project ORM instance.
        2. Add to the database, commit, and return the refreshed object.
    """
    project = Project(**body.model_dump())
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@router.get("/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, db: DB, _: Auth):
    """GET /projects/{project_id} — Retrieve a single project by ID.

    Path Parameters:
        project_id (int): The primary key of the project.

    Returns:
        ProjectOut object, or HTTP 404 if not found.
    """
    return _get_project_or_404(db, project_id)


@router.patch("/{project_id}", response_model=ProjectOut)
def update_project(project_id: int, body: ProjectUpdate, db: DB, _: Auth):
    """PATCH /projects/{project_id} — Partially update a project.

    Path Parameters:
        project_id (int): The primary key of the project to update.

    Request Body:
        ProjectUpdate schema — only fields present in the request body are updated.

    Returns:
        The updated ProjectOut object, or HTTP 404 if not found.

    Business Logic:
        1. Fetch the project or raise 404.
        2. Apply only the fields that were explicitly set in the request (exclude_unset).
        3. Commit and return the refreshed project.
    """
    p = _get_project_or_404(db, project_id)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(p, field, value)
    db.commit()
    db.refresh(p)
    return p


@router.delete("/{project_id}", status_code=204)
def delete_project(project_id: int, db: DB, _: Auth):
    """DELETE /projects/{project_id} — Delete a project.

    Path Parameters:
        project_id (int): The primary key of the project to delete.

    Returns:
        HTTP 204 No Content on success, or HTTP 404 if not found.

    Note:
        Cascade deletes for deliverables, phases, assignments, etc. are
        handled by the database foreign key constraints.
    """
    p = _get_project_or_404(db, project_id)
    db.delete(p)
    db.commit()


# ── Project Consultants ────────────────────────────────────────────────────────

@router.post("/{project_id}/consultants/{consultant_id}", status_code=204)
def add_project_consultant(project_id: int, consultant_id: int, db: DB, _: Auth):
    """POST /projects/{project_id}/consultants/{consultant_id} — Assign a consultant to a project.

    Path Parameters:
        project_id (int): The project to add the consultant to.
        consultant_id (int): The consultant to assign.

    Returns:
        HTTP 204 No Content on success.

    Business Logic:
        1. Validate both the project and consultant exist (404 if not).
        2. If the consultant is not already in the project's team, append and commit.
        3. Idempotent — no error if consultant is already assigned.
    """
    p = _get_project_or_404(db, project_id)
    c = db.get(Consultant, consultant_id)
    if not c:
        raise HTTPException(status_code=404, detail="Consultant not found")
    if c not in p.consultants:
        p.consultants.append(c)
        db.commit()


@router.delete("/{project_id}/consultants/{consultant_id}", status_code=204)
def remove_project_consultant(project_id: int, consultant_id: int, db: DB, _: Auth):
    """DELETE /projects/{project_id}/consultants/{consultant_id} — Remove a consultant from a project.

    Path Parameters:
        project_id (int): The project to remove the consultant from.
        consultant_id (int): The consultant to unassign.

    Returns:
        HTTP 204 No Content on success.

    Business Logic:
        Silently succeeds if the consultant is not found or not assigned — no error raised.
    """
    p = _get_project_or_404(db, project_id)
    c = db.get(Consultant, consultant_id)
    if c and c in p.consultants:
        p.consultants.remove(c)
        db.commit()


# ── Deliverables ──────────────────────────────────────────────────────────────

@router.patch("/{project_id}/deliverables/reorder")
def reorder_deliverables(project_id: int, body: DeliverableReorder, db: DB, _: Auth):
    """PATCH /projects/{project_id}/deliverables/reorder — Reorder deliverables by drag-and-drop.

    Path Parameters:
        project_id (int): The project containing the deliverables.

    Request Body:
        DeliverableReorder schema with ``ids`` — an ordered list of deliverable IDs.

    Returns:
        {"ok": True} on success.

    Business Logic:
        1. Validate the project exists.
        2. Build a mapping of deliverable ID -> new sort_order (based on list index).
        3. Fetch all deliverables matching the IDs in a single query.
        4. Update each deliverable's sort_order and commit.
    """
    _get_project_or_404(db, project_id)
    # Single query instead of N individual lookups
    id_to_order = {did: i for i, did in enumerate(body.ids)}
    deliverables = (
        db.query(Deliverable)
        .filter(Deliverable.id.in_(body.ids), Deliverable.project_id == project_id)
        .all()
    )
    for d in deliverables:
        d.sort_order = id_to_order[d.id]
    db.commit()
    return {"ok": True}


@router.get("/{project_id}/deliverables", response_model=list[DeliverableOut])
def list_deliverables(project_id: int, db: DB, _: Auth):
    """GET /projects/{project_id}/deliverables — List all deliverables for a project.

    Path Parameters:
        project_id (int): The project whose deliverables to list.

    Returns:
        List of DeliverableOut objects ordered by sort_order, with phases
        and their assignments eagerly loaded to avoid N+1 queries.
    """
    _get_project_or_404(db, project_id)
    return (
        db.query(Deliverable)
        .options(selectinload(Deliverable.phases).selectinload(DeliverablePhase.assignments))
        .filter(Deliverable.project_id == project_id)
        .order_by(Deliverable.sort_order)
        .all()
    )


@router.post("/{project_id}/deliverables", response_model=DeliverableOut, status_code=201)
def create_deliverable(project_id: int, body: DeliverableCreate, db: DB, _: Auth):
    """POST /projects/{project_id}/deliverables — Create a new deliverable.

    Path Parameters:
        project_id (int): The project to create the deliverable in.

    Request Body:
        DeliverableCreate schema with name, type, dates, hours, etc.

    Returns:
        The newly created DeliverableOut object (HTTP 201).

    Business Logic:
        1. Validate the project exists.
        2. Create the Deliverable record.
        3. Auto-calculate missing dates: if business_days is set with only start_date,
           compute end_date; if only end_date, compute start_date.
        4. Flush to get the deliverable ID.
        5. Auto-create standard phases (workshop, draft, qa, delivery).
           Workshop phase is skipped for non-workshop deliverable types.
        6. Wire up ORM relationships so sync can traverse phase -> deliverable.
        7. Call sync_deliverable_assignments to create/update WeeklyAllocation rows.
        8. Commit and return the refreshed deliverable.
    """
    project = _get_project_or_404(db, project_id)
    data = body.model_dump()
    d = Deliverable(project_id=project_id, **data)
    # Auto-calculate dates from business_days when one date is missing
    if d.business_days and d.business_days > 0:
        if d.start_date and not d.end_date:
            d.end_date = _add_business_days(d.start_date, d.business_days)
        elif d.end_date and not d.start_date:
            d.start_date = _sub_business_days(d.end_date, d.business_days)
    db.add(d)
    db.flush()  # get d.id before creating phases
    # Auto-create standard phases (skip workshop phase for non-workshop deliverables)
    phase_types = _PHASE_TYPES if body.deliverable_type == "workshop" else _PHASE_TYPES[1:]
    phases = []
    for i, pt in enumerate(phase_types):
        p = DeliverablePhase(deliverable_id=d.id, phase_type=pt, sort_order=i)
        db.add(p)
        phases.append(p)
    db.flush()
    # Wire up relationships so the sync can traverse phase -> deliverable
    d.phases = phases
    for p in phases:
        p.deliverable = d
    sync_deliverable_assignments(db, d, snap_to_friday=project.snap_end_to_friday)
    db.commit()
    db.refresh(d)
    return d


@router.get("/{project_id}/deliverables/{deliverable_id}", response_model=DeliverableOut)
def get_deliverable(project_id: int, deliverable_id: int, db: DB, _: Auth):
    """GET /projects/{project_id}/deliverables/{deliverable_id} — Get a single deliverable.

    Path Parameters:
        project_id (int): The parent project ID.
        deliverable_id (int): The deliverable to retrieve.

    Returns:
        DeliverableOut object, or HTTP 404 if not found within the given project.
    """
    d = db.query(Deliverable).filter(
        Deliverable.id == deliverable_id, Deliverable.project_id == project_id
    ).first()
    if not d:
        raise HTTPException(status_code=404, detail="Deliverable not found")
    return d


@router.patch("/{project_id}/deliverables/{deliverable_id}", response_model=DeliverableOut)
def update_deliverable(project_id: int, deliverable_id: int, body: DeliverableUpdate, db: DB, _: Auth):
    """PATCH /projects/{project_id}/deliverables/{deliverable_id} — Partially update a deliverable.

    Path Parameters:
        project_id (int): The parent project ID.
        deliverable_id (int): The deliverable to update.

    Request Body:
        DeliverableUpdate schema — only fields present in the request body are updated.

    Returns:
        The updated DeliverableOut object.

    Business Logic:
        1. Fetch the project and deliverable (with phases eagerly loaded).
        2. Apply all provided fields to the deliverable.
        3. Auto-calculate dates from business_days using these rules:
           - If start_date was updated and end_date wasn't: compute end_date.
           - If end_date was updated and start_date wasn't: compute start_date.
           - If business_days was updated: recompute from whichever anchor date exists.
        4. Propagate draft_start_date, draft_end_date, qa_start_date, qa_end_date
           to the corresponding phase records.
        5. Re-sync assignments and WeeklyAllocation rows.
        6. Commit and return the refreshed deliverable.
    """
    project = _get_project_or_404(db, project_id)
    d = db.query(Deliverable).options(
        selectinload(Deliverable.phases)
    ).filter(
        Deliverable.id == deliverable_id, Deliverable.project_id == project_id
    ).first()
    if not d:
        raise HTTPException(status_code=404, detail="Deliverable not found")
    updates = body.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(d, field, value)
    # Auto-calculate dates from business_days
    sd = d.start_date
    ed = d.end_date
    bd = d.business_days
    if bd and bd > 0:
        if 'start_date' in updates and sd and (not ed or 'end_date' not in updates):
            d.end_date = _add_business_days(sd, bd)
        elif 'end_date' in updates and ed and (not sd or 'start_date' not in updates):
            d.start_date = _sub_business_days(ed, bd)
        elif 'business_days' in updates and sd:
            d.end_date = _add_business_days(sd, bd)
        elif 'business_days' in updates and ed:
            d.start_date = _sub_business_days(ed, bd)
    # Handle phase date updates — propagate convenience fields to actual phase records
    draft_phase = next((p for p in d.phases if p.phase_type == "draft"), None)
    qa_phase = next((p for p in d.phases if p.phase_type == "qa"), None)
    if 'draft_start_date' in updates and draft_phase:
        draft_phase.start_date = updates['draft_start_date']
    if 'draft_end_date' in updates and draft_phase:
        draft_phase.end_date = updates['draft_end_date']
    if 'qa_start_date' in updates and qa_phase:
        qa_phase.start_date = updates['qa_start_date']
    if 'qa_end_date' in updates and qa_phase:
        qa_phase.end_date = updates['qa_end_date']
    # Ensure phase -> deliverable back-references are wired for sync
    for p in d.phases:
        p.deliverable = d
    sync_deliverable_assignments(db, d, snap_to_friday=project.snap_end_to_friday)
    db.commit()
    db.refresh(d)
    return d


@router.post("/{project_id}/sync-allocations")
def sync_project_allocations(project_id: int, db: DB, _: Auth):
    """POST /projects/{project_id}/sync-allocations — Re-sync all allocations for a project.

    Backfill: re-sync assignments and WeeklyAllocations for every deliverable in the project.
    Safe to call multiple times — idempotent.

    Path Parameters:
        project_id (int): The project to sync.

    Returns:
        {"synced": <count>} indicating how many deliverables were processed.

    Business Logic:
        1. Load all deliverables with their phases.
        2. For each deliverable, wire up phase -> deliverable back-references.
        3. Call sync_deliverable_assignments for each deliverable.
        4. Commit once after all deliverables are processed.
    """
    project = _get_project_or_404(db, project_id)
    deliverables = (
        db.query(Deliverable)
        .options(selectinload(Deliverable.phases))
        .filter(Deliverable.project_id == project_id)
        .all()
    )
    for d in deliverables:
        for p in d.phases:
            p.deliverable = d
        sync_deliverable_assignments(db, d, snap_to_friday=project.snap_end_to_friday)
    db.commit()
    return {"synced": len(deliverables)}




def _next_business_day(d: dt_date) -> dt_date:
    """Return the next business day after d (i.e., add exactly 1 business day)."""
    return _add_business_days(d, 1)


def _prev_business_day(d: dt_date) -> dt_date:
    """Return the previous business day before d (i.e., subtract exactly 1 business day)."""
    return _sub_business_days(d, 1)


@router.post("/{project_id}/deliverables/{deliverable_id}/auto-dates", response_model=DeliverableOut)
def auto_generate_dates(project_id: int, deliverable_id: int, db: DB, _: Auth):
    """POST /projects/{project_id}/deliverables/{deliverable_id}/auto-dates — Auto-generate phase dates.

    Auto-generate all dates for a deliverable based on rules:
    - consultant_business_days and qa_business_days must be set
    - Both workshop_sequential and qa_sequential must be true
    - Either start_date or end_date must be set
    Forward from start_date or backward from end_date.

    Path Parameters:
        project_id (int): The parent project ID.
        deliverable_id (int): The deliverable to auto-date.

    Returns:
        The updated DeliverableOut with all dates populated.

    Business Logic (Forward — when start_date is set):
        1. Determine consultant start: use workshop date + 1 business day if workshop_sequential,
           otherwise use the deliverable's start_date.
        2. Consultant end = cons_start + consultant_business_days.
        3. QA start = next business day after cons_end if qa_sequential, else cons_end.
        4. QA end = qa_start + qa_business_days.
        5. Set deliverable end_date = qa_end.
        6. Update draft and QA phase records with computed dates and hours.

    Business Logic (Backward — when only end_date is set):
        1. QA end = deliverable end_date.
        2. QA start = qa_end - qa_business_days.
        3. Consultant end = prev business day before qa_start if qa_sequential, else qa_start.
        4. Consultant start = cons_end - consultant_business_days.
        5. Set deliverable start_date = workshop date (if sequential) or cons_start.
        6. Update draft and QA phase records.
    """
    project = _get_project_or_404(db, project_id)
    d = db.query(Deliverable).options(
        selectinload(Deliverable.phases)
    ).filter(
        Deliverable.id == deliverable_id, Deliverable.project_id == project_id
    ).first()
    if not d:
        raise HTTPException(status_code=404, detail="Deliverable not found")

    cbd = d.business_days
    qbd = d.qa_business_days
    if not cbd or cbd <= 0:
        raise HTTPException(status_code=400, detail="Consultant Business Days must be set")
    if not qbd or qbd <= 0:
        raise HTTPException(status_code=400, detail="QA Business Days must be set")

    draft_phase = next((p for p in d.phases if p.phase_type == "draft"), None)
    qa_phase = next((p for p in d.phases if p.phase_type == "qa"), None)

    # Get workshop date if linked and sequential
    ws_date = None
    if d.workshop_sequential and d.workshop_id:
        ws = db.get(Workshop, d.workshop_id)
        if ws and ws.workshop_date:
            ws_date = ws.workshop_date

    if d.start_date:
        # Forward generation: compute dates from start_date forward
        cons_start = d.start_date
        if ws_date and d.workshop_sequential:
            # Consultant work begins the business day after the workshop
            cons_start = _next_business_day(ws_date)
            if cons_start < d.start_date:
                cons_start = d.start_date

        cons_end = _add_business_days(cons_start, cbd)
        # QA starts after consultant work; next day if sequential, same day otherwise
        qa_start = _next_business_day(cons_end) if d.qa_sequential else cons_end
        qa_end = _add_business_days(qa_start, qbd)

        d.end_date = qa_end
        if draft_phase:
            draft_phase.start_date = cons_start
            draft_phase.end_date = cons_end
            draft_phase.total_hours = float(d.flat_hours or 0)
        if qa_phase:
            qa_phase.start_date = qa_start
            qa_phase.end_date = qa_end
            qa_phase.total_hours = float(d.qa_hours or 0)

    elif d.end_date:
        # Backward generation: compute dates from end_date backward
        qa_end = d.end_date
        qa_start = _sub_business_days(qa_end, qbd)
        # Consultant ends before QA starts; prev day if sequential, same day otherwise
        cons_end = _prev_business_day(qa_start) if d.qa_sequential else qa_start
        cons_start = _sub_business_days(cons_end, cbd)

        if ws_date and d.workshop_sequential:
            d.start_date = ws_date
        else:
            d.start_date = cons_start

        if draft_phase:
            draft_phase.start_date = cons_start
            draft_phase.end_date = cons_end
            draft_phase.total_hours = float(d.flat_hours or 0)
        if qa_phase:
            qa_phase.start_date = qa_start
            qa_phase.end_date = qa_end
            qa_phase.total_hours = float(d.qa_hours or 0)
    else:
        raise HTTPException(status_code=400, detail="Either Start Date or Due Date must be set")

    # Wire up phase -> deliverable back-references and sync allocations
    for p in d.phases:
        p.deliverable = d
    sync_deliverable_assignments(db, d, snap_to_friday=project.snap_end_to_friday)
    db.commit()
    db.refresh(d)
    return d

@router.delete("/{project_id}/deliverables/{deliverable_id}", status_code=204)
def delete_deliverable(project_id: int, deliverable_id: int, db: DB, _: Auth):
    """DELETE /projects/{project_id}/deliverables/{deliverable_id} — Delete a deliverable.

    Path Parameters:
        project_id (int): The parent project ID.
        deliverable_id (int): The deliverable to delete.

    Returns:
        HTTP 204 No Content on success, or HTTP 404 if not found.

    Note:
        Associated phases, assignments, and weekly allocations are cascade-deleted.
    """
    d = db.query(Deliverable).filter(
        Deliverable.id == deliverable_id, Deliverable.project_id == project_id
    ).first()
    if not d:
        raise HTTPException(status_code=404, detail="Deliverable not found")
    db.delete(d)
    db.commit()


# ── Phases ────────────────────────────────────────────────────────────────────

@router.patch("/{project_id}/deliverables/{deliverable_id}/phases/{phase_id}", response_model=PhaseOut)
def update_phase(project_id: int, deliverable_id: int, phase_id: int, body: PhaseUpdate, db: DB, _: Auth):
    """PATCH /projects/{project_id}/deliverables/{deliverable_id}/phases/{phase_id} — Update a phase.

    Path Parameters:
        project_id (int): The parent project ID.
        deliverable_id (int): The parent deliverable ID.
        phase_id (int): The phase to update.

    Request Body:
        PhaseUpdate schema — supports updating start_date, end_date, total_hours, etc.

    Returns:
        The updated PhaseOut object, or HTTP 404 if the phase is not found
        within the specified deliverable and project.

    Business Logic:
        1. Join DeliverablePhase with Deliverable to verify the phase belongs
           to the correct deliverable and project.
        2. Apply only the provided fields.
        3. Commit and return.
    """
    phase = db.query(DeliverablePhase).join(Deliverable).filter(
        DeliverablePhase.id == phase_id,
        Deliverable.id == deliverable_id,
        Deliverable.project_id == project_id,
    ).first()
    if not phase:
        raise HTTPException(status_code=404, detail="Phase not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(phase, field, value)
    db.commit()
    db.refresh(phase)
    return phase


# ── Workshops ─────────────────────────────────────────────────────────────────

@router.get("/{project_id}/workshops", response_model=list[WorkshopOut])
def list_workshops(project_id: int, db: DB, _: Auth):
    """GET /projects/{project_id}/workshops — List all workshops for a project.

    Path Parameters:
        project_id (int): The project whose workshops to list.

    Returns:
        List of WorkshopOut objects ordered by workshop_date.
    """
    _get_project_or_404(db, project_id)
    return db.query(Workshop).filter(Workshop.project_id == project_id).order_by(Workshop.workshop_date).all()


@router.post("/{project_id}/workshops", response_model=WorkshopOut, status_code=201)
def create_workshop(project_id: int, body: WorkshopCreate, db: DB, _: Auth):
    """POST /projects/{project_id}/workshops — Create a new workshop.

    Path Parameters:
        project_id (int): The project to create the workshop in.

    Request Body:
        WorkshopCreate schema with name, workshop_date, duration_hours,
        and optional consultant_ids to assign.

    Returns:
        The newly created WorkshopOut object (HTTP 201).

    Business Logic:
        1. Validate the project exists.
        2. Create the Workshop record (excluding consultant_ids from model_dump).
        3. If consultant_ids are provided, look up the Consultant records and
           assign them to the workshop via the many-to-many relationship.
        4. Commit and return.
    """
    _get_project_or_404(db, project_id)
    data = body.model_dump(exclude={"consultant_ids"})
    w = Workshop(project_id=project_id, **data)
    if body.consultant_ids:
        w.consultants = db.query(Consultant).filter(Consultant.id.in_(body.consultant_ids)).all()
    db.add(w)
    db.commit()
    db.refresh(w)
    return w


@router.patch("/{project_id}/workshops/{workshop_id}", response_model=WorkshopOut)
def update_workshop(project_id: int, workshop_id: int, body: WorkshopUpdate, db: DB, _: Auth):
    """PATCH /projects/{project_id}/workshops/{workshop_id} — Update a workshop.

    Path Parameters:
        project_id (int): The parent project ID.
        workshop_id (int): The workshop to update.

    Request Body:
        WorkshopUpdate schema — partial update of name, date, duration, consultant_ids.

    Returns:
        The updated WorkshopOut object, or HTTP 404 if not found.

    Business Logic:
        1. Fetch the workshop or raise 404.
        2. Apply non-consultant fields.
        3. If consultant_ids is explicitly provided (even as empty list),
           replace the entire consultant list.
        4. Commit and return.
    """
    w = db.query(Workshop).filter(Workshop.id == workshop_id, Workshop.project_id == project_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Workshop not found")
    for field, value in body.model_dump(exclude_unset=True, exclude={"consultant_ids"}).items():
        setattr(w, field, value)
    if body.consultant_ids is not None:
        w.consultants = db.query(Consultant).filter(Consultant.id.in_(body.consultant_ids)).all()
    db.commit()
    db.refresh(w)
    return w


@router.delete("/{project_id}/workshops/{workshop_id}", status_code=204)
def delete_workshop(project_id: int, workshop_id: int, db: DB, _: Auth):
    """DELETE /projects/{project_id}/workshops/{workshop_id} — Delete a workshop.

    Path Parameters:
        project_id (int): The parent project ID.
        workshop_id (int): The workshop to delete.

    Returns:
        HTTP 204 No Content on success, or HTTP 404 if not found.
    """
    w = db.query(Workshop).filter(Workshop.id == workshop_id, Workshop.project_id == project_id).first()
    if not w:
        raise HTTPException(status_code=404, detail="Workshop not found")
    db.delete(w)
    db.commit()


# ── Export Project to Excel ────────────────────────────────────────────────────

import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@router.get('/{project_id}/export-excel')
def export_project_excel(project_id: int, db: DB, _: Auth):
    """GET /projects/{project_id}/export-excel — Export a project as a multi-sheet Excel workbook.

    Path Parameters:
        project_id (int): The project to export.

    Returns:
        A streaming .xlsx file download with three sheets:
        1. **Project Overview** — project metadata, team members, hours by consultant.
        2. **Deliverables** — all deliverables with dates, hours, phase info, and totals.
        3. **Workshops** — all workshops with dates, durations, consultants, and totals.

    Business Logic:
        1. Fetch the project, its deliverables (with phases), and workshops.
        2. Build a consultant name lookup map.
        3. Create styled Excel sheets with headers, data rows, and summary totals.
        4. Calculate aggregate hours (deliverable hours, workshop hours, per-consultant breakdown).
        5. Stream the workbook as a downloadable .xlsx file.
    """
    project = _get_project_or_404(db, project_id)
    deliverables = (
        db.query(Deliverable)
        .options(selectinload(Deliverable.phases))
        .filter(Deliverable.project_id == project_id)
        .order_by(Deliverable.sort_order)
        .all()
    )
    workshops = (
        db.query(Workshop)
        .filter(Workshop.project_id == project_id)
        .all()
    )
    from app.models.consultant import Consultant
    # Build a map of consultant ID -> name for display in the spreadsheet
    consultants_map = {c.id: c.name for c in db.query(Consultant).all()}

    wb = Workbook()
    # Define reusable styles for headers, sections, and borders
    hdr_font = Font(bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    section_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def styled_header(ws, row, headers):
        """Apply header styling (white text on blue background) to a row of cells."""
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin

    def data_cell(ws, row, col, value):
        """Write a value to a cell with thin borders applied."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin
        return cell

    # ─── Project Overview sheet ──────────────────────────────────
    ov = wb.active
    ov.title = 'Project Overview'
    ov.sheet_properties.tabColor = '2563EB'

    # Calculate aggregate hours across all deliverables and workshops
    total_deliv_hrs = sum(float(d.flat_hours or 0) + float(d.qa_hours or 0) for d in deliverables)
    total_ws_hrs = sum(
        float(w.duration_hours or 0) * len(w.consultants)
        for w in workshops
    )

    # Key-value pairs for the project overview section
    info_fields = [
        ('Project Name', project.name),
        ('Client', project.client_name or ''),
        ('Status', (project.status or '').replace('_', ' ').title()),
        ('Start Date', str(project.start_date) if project.start_date else ''),
        ('End Date', str(project.end_date) if project.end_date else ''),
        ('Budgeted Hours', project.budgeted_hours),
        ('Total Planned Hours', total_deliv_hrs + total_ws_hrs),
        ('Deliverable Hours', total_deliv_hrs),
        ('Workshop Hours', total_ws_hrs),
        ('Deliverable Count', len(deliverables)),
        ('Workshop Count', len(workshops)),
        ('Notes', project.notes or ''),
    ]

    styled_header(ov, 1, ['Field', 'Value'])
    for i, (field, val) in enumerate(info_fields, 2):
        c1 = data_cell(ov, i, 1, field)
        c1.font = Font(bold=True)
        data_cell(ov, i, 2, val)

    # Team Members section — lists all consultants assigned to the project
    team_row = len(info_fields) + 4
    ov.cell(row=team_row, column=1, value='Team Members').font = Font(bold=True, size=12, color='2563EB')
    styled_header(ov, team_row + 1, ['Name', 'Role'])
    for i, c in enumerate(project.consultants, team_row + 2):
        data_cell(ov, i, 1, c.name)
        data_cell(ov, i, 2, 'Consultant')

    # Hours by consultant — aggregate deliverable and workshop hours per person
    con_hrs = {}
    for d in deliverables:
        if d.consultant_id and d.flat_hours:
            con_hrs[d.consultant_id] = con_hrs.get(d.consultant_id, 0) + float(d.flat_hours)
        if d.qa_consultant_id and d.qa_hours:
            con_hrs[d.qa_consultant_id] = con_hrs.get(d.qa_consultant_id, 0) + float(d.qa_hours)
    for w in workshops:
        wh = float(w.duration_hours or 0)
        for wc in w.consultants:
            con_hrs[wc.id] = con_hrs.get(wc.id, 0) + wh

    hrs_row = team_row + len(project.consultants) + 4
    ov.cell(row=hrs_row, column=1, value='Hours by Consultant').font = Font(bold=True, size=12, color='2563EB')
    styled_header(ov, hrs_row + 1, ['Consultant', 'Deliverable Hrs', 'Workshop Hrs', 'Total Hrs'])
    row = hrs_row + 2
    for cid, hrs in sorted(con_hrs.items(), key=lambda x: -x[1]):
        cname = consultants_map.get(cid, 'Unknown')
        # Split deliverable vs workshop hours for each consultant
        d_hrs = 0
        for d in deliverables:
            if d.consultant_id == cid:
                d_hrs += float(d.flat_hours or 0)
            if d.qa_consultant_id == cid:
                d_hrs += float(d.qa_hours or 0)
        w_hrs = 0
        for w in workshops:
            if cid in [wc.id for wc in w.consultants]:
                w_hrs += float(w.duration_hours or 0)
        data_cell(ov, row, 1, cname)
        data_cell(ov, row, 2, d_hrs)
        data_cell(ov, row, 3, w_hrs)
        data_cell(ov, row, 4, d_hrs + w_hrs)
        row += 1

    # Set column widths for the overview sheet
    ov.column_dimensions['A'].width = 25
    ov.column_dimensions['B'].width = 35
    ov.column_dimensions['C'].width = 15
    ov.column_dimensions['D'].width = 12

    # ─── Deliverables sheet ──────────────────────────────────────
    ds = wb.create_sheet('Deliverables')
    ds.sheet_properties.tabColor = 'EA580C'
    d_headers = [
        'Name', 'Type', 'Status', 'Consultant', 'QA Consultant',
        'Start Date', 'Due Date', 'Consultant Days', 'QA Days',
        'Consultant Hours', 'QA Hours', 'Total Hours',
        'Consultant Due', 'QA Due',
        'Workshop Sequential', 'QA Sequential',
    ]
    styled_header(ds, 1, d_headers)

    # Write one row per deliverable with all key fields
    for i, d in enumerate(deliverables, 2):
        draft_phase = next((p for p in d.phases if p.phase_type == 'draft'), None)
        qa_phase = next((p for p in d.phases if p.phase_type == 'qa'), None)
        vals = [
            d.name,
            d.deliverable_type,
            (d.status or '').replace('_', ' ').title(),
            consultants_map.get(d.consultant_id, '') if d.consultant_id else '',
            consultants_map.get(d.qa_consultant_id, '') if d.qa_consultant_id else '',
            str(d.start_date) if d.start_date else '',
            str(d.end_date) if d.end_date else '',
            d.business_days,
            d.qa_business_days,
            float(d.flat_hours) if d.flat_hours else None,
            float(d.qa_hours) if d.qa_hours else None,
            d.total_planned_hours,
            str(draft_phase.end_date) if draft_phase and draft_phase.end_date else '',
            str(qa_phase.end_date) if qa_phase and qa_phase.end_date else '',
            'Yes' if d.workshop_sequential else 'No',
            'Yes' if d.qa_sequential else 'No',
        ]
        for col, val in enumerate(vals, 1):
            data_cell(ds, i, col, val)

    # Totals row at the bottom of the deliverables sheet
    total_row = len(deliverables) + 2
    ds.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
    data_cell(ds, total_row, 10, sum(float(d.flat_hours or 0) for d in deliverables))
    data_cell(ds, total_row, 11, sum(float(d.qa_hours or 0) for d in deliverables))
    data_cell(ds, total_row, 12, total_deliv_hrs)
    for col in [10, 11, 12]:
        ds.cell(row=total_row, column=col).font = Font(bold=True)

    col_widths = [35, 14, 14, 18, 18, 12, 12, 14, 10, 16, 12, 12, 14, 12, 18, 14]
    for col, w in enumerate(col_widths, 1):
        ds.column_dimensions[get_column_letter(col)].width = w

    # ─── Workshops sheet ─────────────────────────────────────────
    ws_sheet = wb.create_sheet('Workshops')
    ws_sheet.sheet_properties.tabColor = '7C3AED'
    w_headers = ['Name', 'Date', 'Duration (hrs)', 'Status', 'Consultants', 'Planned Hours']
    styled_header(ws_sheet, 1, w_headers)

    # Write one row per workshop
    for i, w in enumerate(workshops, 2):
        cons_names = ', '.join(c.name for c in w.consultants)
        # Planned hours = duration * number of consultants attending
        planned = float(w.duration_hours or 0) * len(w.consultants)
        vals = [
            w.name,
            str(w.workshop_date) if w.workshop_date else '',
            float(w.duration_hours) if w.duration_hours else None,
            (w.status or '').replace('_', ' ').title(),
            cons_names,
            planned,
        ]
        for col, val in enumerate(vals, 1):
            data_cell(ws_sheet, i, col, val)

    # Totals row for workshops
    total_ws_row = len(workshops) + 2
    ws_sheet.cell(row=total_ws_row, column=1, value='TOTALS').font = Font(bold=True)
    data_cell(ws_sheet, total_ws_row, 6, total_ws_hrs)
    ws_sheet.cell(row=total_ws_row, column=6).font = Font(bold=True)

    ws_sheet.column_dimensions['A'].width = 35
    ws_sheet.column_dimensions['B'].width = 14
    ws_sheet.column_dimensions['C'].width = 16
    ws_sheet.column_dimensions['D'].width = 16
    ws_sheet.column_dimensions['E'].width = 40
    ws_sheet.column_dimensions['F'].width = 14

    # Serialize workbook to an in-memory buffer and stream as a download
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Sanitize the project name for use as a filename
    safe_name = project.name.replace(' ', '_').replace('/', '-')[:50]
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={safe_name}_export.xlsx'},
    )
