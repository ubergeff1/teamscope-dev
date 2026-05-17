"""
Project templates router — CRUD for project templates, deliverable templates, and workshop templates.

API prefix: /api/templates
Tags: ["templates"]

Templates are blueprints that pre-populate deliverables and workshops when creating
a new project. This router supports:
  - Downloading a blank Excel template file with instructions for offline editing
  - Importing a filled-in Excel file to create a template with deliverables and workshops
  - Full CRUD for project templates, deliverable templates, and workshop templates
  - Applying a template to an existing project (cloning deliverables/workshops)

Key endpoint:
  POST /api/templates/apply/{project_id}/{template_id} — applies a template to a project.
"""
import io
from typing import Annotated
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.template import ProjectTemplate, DeliverableTemplate, WorkshopTemplate
from app.models.reference import ControlFamily
from app.models.deliverable import Deliverable, DeliverablePhase, Workshop
from app.models.project import Project
from app.schemas.template import (
    ProjectTemplateCreate,
    ProjectTemplateUpdate,
    ProjectTemplateOut,
    DeliverableTemplateCreate,
    DeliverableTemplateUpdate,
    DeliverableTemplateOut,
    WorkshopTemplateCreate,
    WorkshopTemplateUpdate,
    WorkshopTemplateOut,
)
from app.utils.auth import get_current_user

# Create the FastAPI router with /templates prefix
router = APIRouter(prefix="/templates", tags=["templates"])

# Type aliases for dependency injection
# Auth: extracts the authenticated username from the JWT token
Auth = Annotated[str, Depends(get_current_user)]
# DB: provides a SQLAlchemy database session
DB = Annotated[Session, Depends(get_db)]

# Standard phase types created for each deliverable when a template is applied.
# "workshop" phase is only included for workshop-type deliverables.
_PHASE_TYPES = ["workshop", "draft", "qa", "delivery"]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _load_template(db: Session, template_id: int) -> ProjectTemplate | None:
    """Fetch a project template by ID with deliverable and workshop templates eagerly loaded.

    Uses selectinload to avoid N+1 queries when accessing the template's
    child deliverable_templates and workshop_templates collections.

    Args:
        db: The database session.
        template_id: The primary key of the project template.

    Returns:
        The ProjectTemplate instance with relationships loaded, or None if not found.
    """
    return (
        db.query(ProjectTemplate)
        .options(
            selectinload(ProjectTemplate.deliverable_templates),
            selectinload(ProjectTemplate.workshop_templates),
        )
        .filter(ProjectTemplate.id == template_id)
        .first()
    )


def _safe_float(val):
    """Safely convert a value to float, returning None on failure.

    Used when parsing Excel cells that may contain empty strings,
    None values, or non-numeric text.

    Args:
        val: The value to convert (from an Excel cell).

    Returns:
        The float value, or None if conversion fails.
    """
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val):
    """Safely convert a value to int (via float to handle decimals), returning None on failure.

    Used when parsing Excel cells for integer fields like business_days.

    Args:
        val: The value to convert (from an Excel cell).

    Returns:
        The int value, or None if conversion fails.
    """
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _styled_header(ws, col, value, font, fill, border):
    """Apply header styling to a single cell in row 1 of a worksheet.

    Args:
        ws: The openpyxl worksheet.
        col: Column number (1-indexed).
        value: The header text.
        font: Font style to apply.
        fill: Fill/background style to apply.
        border: Border style to apply.

    Returns:
        The styled cell object.
    """
    cell = ws.cell(row=1, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border
    return cell


def _note_cell(ws, row, col, value):
    """Write a note/instruction cell with italic gray styling.

    Used to add help text below data rows in template Excel files.

    Args:
        ws: The openpyxl worksheet.
        row: Row number (1-indexed).
        col: Column number (1-indexed).
        value: The note text.

    Returns:
        The styled cell object.
    """
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(italic=True, color="666666", size=10)
    return cell


# ── Excel Template Download ────────────────────────────────────────────────────

@router.get("/excel-template")
def download_excel_template(_: Auth):
    """GET /templates/excel-template — Download a blank Excel template for offline editing.

    Returns a formatted .xlsx workbook with four sheets:
      1. **Instructions** — detailed guide on how to fill in the template
      2. **Template Info** — name and description fields for the template
      3. **Workshops** — columns for defining workshop blueprints (name, duration)
      4. **Deliverables** — columns for defining deliverable blueprints
         (name, type, control family code, hours, business days, workshop link)

    Each data sheet includes example rows (prefixed with "Example:") that are
    automatically ignored during import. The user fills in real data, then uploads
    the file via the POST /templates/import-excel endpoint.

    Returns:
        A streaming .xlsx file download named "template_import.xlsx".
    """
    wb = Workbook()

    # Define reusable styles for the entire workbook
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    section_font = Font(bold=True, size=11, color="2563EB")
    note_font = Font(italic=True, color="666666", size=10)
    example_font = Font(italic=True, color="AAAAAA")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ─── Instructions sheet ───────────────────────────────────────────────
    # Provides a comprehensive guide for filling in the template
    inst = wb.active
    inst.title = "Instructions"
    inst.sheet_properties.tabColor = "2563EB"
    inst.column_dimensions["A"].width = 90

    rows = [
        ("TEAMSCOPE TEMPLATE IMPORT — INSTRUCTIONS", Font(bold=True, size=14, color="2563EB")),
        ("", None),
        ("This workbook has 4 sheets. Fill in the sheets described below, then upload this file", note_font),
        ("using the 'Import Template' button on the Templates page.", note_font),
        ("Rows that start with 'Example:' are ignored on import — delete or leave them.", note_font),
        ("", None),
        ("═══ SHEET: Template Info ═══", section_font),
        ("  Template Name  (required)  — The name for your template, e.g. 'CMMC Level 2'", None),
        ("  Description    (optional)  — A short description of what this template covers", None),
        ("", None),
        ("═══ SHEET: Workshops ═══", section_font),
        ("  Define workshops that will be created when this template is applied to a project.", None),
        ("  Columns:", None),
        ("    Name              (required)  — Workshop name, e.g. 'Kickoff Workshop'", None),
        ("    Duration (hours)  (optional)  — Length of the workshop in hours, e.g. 1.5", None),
        ("", None),
        ("═══ SHEET: Deliverables ═══", section_font),
        ("  Define deliverables that will be created when this template is applied.", None),
        ("  Columns:", None),
        ("    Name                (required)  — Deliverable name", None),
        ("    Type                (required)  — One of the following values:", None),
        ("        control_family  — A NIST/CMMC control family. Set 'Control Family Code' and 'Hours Per Control'.", None),
        ("        flat_hours      — A fixed-hours deliverable. Set 'Flat Hours'.", None),
        ("        appendix        — An appendix document. Set 'Flat Hours'.", None),
        ("        workshop        — A deliverable tied to a workshop phase. Set 'Flat Hours'.", None),
        ("        custom          — Any other deliverable type. Set 'Flat Hours'.", None),
        ("    Control Family Code (conditional) — e.g. 'AC', 'AU', 'CM'. Only for control_family type.", None),
        ("    Hours Per Control   (conditional) — Hours per individual control. Only for control_family type.", None),
        ("    Flat Hours          (optional)  — Total consultant hours for non-control-family deliverables.", None),
        ("    QA Hours            (optional)  — Hours allocated for QA review.", None),
        ("    Business Days       (optional)  — Number of working days to complete.", None),
        ("    Workshop            (optional)  — Name of a workshop from the Workshops sheet to link this", None),
        ("                                      deliverable to. Must match exactly (case-insensitive).", None),
        ("", None),
        ("═══ HOW WORKSHOP LINKING WORKS ═══", section_font),
        ("  Deliverables can be associated with a workshop. When the template is applied:", None),
        ("  1. Workshops are created first", None),
        ("  2. Each deliverable with a 'Workshop' value is linked to the matching workshop", None),
        ("  3. The workshop name in the Deliverables sheet must match a name in the Workshops sheet", None),
        ("", None),
        ("  Example:  Workshops sheet has 'Scoping Workshop'", None),
        ("            Deliverables sheet has a row with Workshop = 'Scoping Workshop'", None),
        ("            → When applied, that deliverable will be linked to the Scoping Workshop", None),
        ("", None),
        ("═══ TIPS ═══", section_font),
        ("  • You can leave optional fields blank — they will default to empty/null", None),
        ("  • Control family deliverables: hours are calculated as control_count × hours_per_control", None),
        ("    (control_count is determined by the project's impact level when the template is applied)", None),
        ("  • Flat hours deliverables: set the total hours directly in the 'Flat Hours' column", None),
        ("  • Sort order is determined by row order in the spreadsheet", None),
    ]

    for i, (text, font) in enumerate(rows, 1):
        cell = inst.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    # ─── Template Info sheet ──────────────────────────────────────────────
    # Two-row key/value sheet for template name and description
    info = wb.create_sheet("Template Info")
    info.sheet_properties.tabColor = "16A34A"
    info["A1"] = "Field"
    info["B1"] = "Value"
    for cell in [info["A1"], info["B1"]]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    fields = [
        ("Template Name", ""),
        ("Description", ""),
    ]
    for row_idx, (field, val) in enumerate(fields, start=2):
        c1 = info.cell(row=row_idx, column=1, value=field)
        c1.border = thin
        c1.font = Font(bold=True)
        c2 = info.cell(row=row_idx, column=2, value=val)
        c2.border = thin

    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 50

    # ─── Workshops sheet ──────────────────────────────────────────────────
    # Two-column sheet: Name and Duration (hours)
    ws_sheet = wb.create_sheet("Workshops")
    ws_sheet.sheet_properties.tabColor = "7C3AED"
    w_headers = ["Name", "Duration (hours)"]
    for col, h in enumerate(w_headers, 1):
        _styled_header(ws_sheet, col, h, hdr_font, hdr_fill, thin)

    # Example rows that are ignored on import (prefixed with "Example:")
    w_examples = [
        ("Example: Kickoff Workshop", 2),
        ("Example: Scoping Workshop", 1.5),
        ("Example: Gap Analysis Review", 1),
    ]
    for row_idx, ex in enumerate(w_examples, start=2):
        for col, val in enumerate(ex, 1):
            cell = ws_sheet.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            cell.font = example_font

    ws_sheet.column_dimensions["A"].width = 35
    ws_sheet.column_dimensions["B"].width = 20

    # ─── Deliverables sheet ───────────────────────────────────────────────
    # Eight-column sheet covering all deliverable template fields
    ds = wb.create_sheet("Deliverables")
    ds.sheet_properties.tabColor = "EA580C"
    d_headers = [
        "Name", "Type", "Control Family Code", "Hours Per Control",
        "Flat Hours", "QA Hours", "Business Days", "Workshop",
    ]
    for col, h in enumerate(d_headers, 1):
        _styled_header(ds, col, h, hdr_font, hdr_fill, thin)

    # Example rows demonstrating different deliverable types
    d_examples = [
        ("Example: AC Control Family", "control_family", "AC", 2.5, None, 4, 10, None),
        ("Example: System Security Plan", "flat_hours", None, None, 40, 8, 15, None),
        ("Example: Scoping Deliverable", "workshop", None, None, 8, 2, 5, "Scoping Workshop"),
    ]
    for row_idx, ex in enumerate(d_examples, start=2):
        for col, val in enumerate(ex, 1):
            cell = ds.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            cell.font = example_font

    # Column widths for readability
    col_widths = [35, 18, 22, 18, 12, 12, 14, 28]
    for col, w in enumerate(col_widths, 1):
        ds.column_dimensions[get_column_letter(col)].width = w

    # Help notes below the example rows
    note_row = len(d_examples) + 3
    notes = [
        "Valid Type values: control_family, flat_hours, appendix, workshop, custom",
        "Workshop column: enter the exact workshop name from the Workshops sheet to link a deliverable to it",
        "Control Family Code: only used when Type = control_family (e.g. AC, AU, CM, IA, SC)",
        "Hours Per Control: only used when Type = control_family",
        "Delete the example rows above and enter your own deliverables",
    ]
    for i, note in enumerate(notes):
        _note_cell(ds, note_row + i, 1, note)

    # Serialize and stream the workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import.xlsx"},
    )


# ── Import Excel ──────────────────────────────────────────────────────────────

@router.post("/import-excel", response_model=ProjectTemplateOut, status_code=201)
def import_excel_template(file: UploadFile = File(...), db: DB = None, _: Auth = None):
    """POST /templates/import-excel — Import an Excel file to create a project template.

    Parse an uploaded Excel file (created from the /excel-template download) and
    create a ProjectTemplate with its deliverable templates and workshop templates.

    Request Body:
        Multipart file upload — the .xlsx file to import.

    Returns:
        The newly created ProjectTemplateOut (HTTP 201).

    Business Logic:
        1. Validate the file is a .xlsx file.
        2. Parse the "Template Info" sheet for template name and description.
        3. Parse the "Workshops" sheet first, building a name -> WorkshopTemplate map
           so that deliverables can reference workshops by name.
        4. Parse the "Deliverables" sheet, creating DeliverableTemplate records.
           - Rows starting with "Example:" are skipped.
           - Workshop linking: column H (Workshop) is matched case-insensitively
             against the workshop name map built in step 3.
           - control_family type: uses Control Family Code and Hours Per Control columns.
           - Other types: use Flat Hours directly.
        5. Commit all records and return the fully loaded template.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx)")

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file. Ensure it is a valid .xlsx file.")

    # --- Parse Template Info sheet ---
    if "Template Info" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Missing 'Template Info' sheet. Download the template first.")

    info = wb["Template Info"]
    # Build a key-value map from the two-column Template Info sheet
    info_map = {}
    for row in info.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            info_map[str(row[0]).strip()] = row[1]

    template_name = str(info_map.get("Template Name", "")).strip()
    if not template_name:
        raise HTTPException(status_code=400, detail="Please provide a Template Name in the 'Template Info' sheet")

    description = str(info_map.get("Description", "") or "").strip()

    # Create the parent template record
    pt = ProjectTemplate(name=template_name, description=description or None)
    db.add(pt)
    db.flush()  # Get pt.id for child records

    # --- Parse Workshops first (so deliverables can reference them by name) ---
    wt_name_map = {}  # lowercase workshop name -> WorkshopTemplate instance
    if "Workshops" in wb.sheetnames:
        ws_sheet = wb["Workshops"]
        for idx, row in enumerate(ws_sheet.iter_rows(min_row=2, max_col=2, values_only=True)):
            name = str(row[0]).strip() if row[0] else ""
            # Skip empty rows and example rows
            if not name or name.startswith("Example:"):
                continue
            duration = _safe_float(row[1])
            wt = WorkshopTemplate(
                project_template_id=pt.id,
                name=name,
                duration_hours=duration,
                sort_order=idx,
            )
            db.add(wt)
            db.flush()  # Get wt.id for deliverable linking
            wt_name_map[name.lower()] = wt

    # --- Parse Deliverables sheet ---
    valid_types = {"control_family", "flat_hours", "appendix", "workshop", "custom"}
    if "Deliverables" in wb.sheetnames:
        d_sheet = wb["Deliverables"]
        for idx, row in enumerate(d_sheet.iter_rows(min_row=2, max_col=8, values_only=True)):
            name = str(row[0]).strip() if row[0] else ""
            # Skip empty rows and example rows
            if not name or name.startswith("Example:"):
                continue
            # Validate and normalize the deliverable type
            dtype = str(row[1]).strip().lower() if row[1] else "flat_hours"
            if dtype not in valid_types:
                dtype = "flat_hours"
            # Parse remaining columns
            cf_code = str(row[2]).strip() if row[2] else None
            hpc = _safe_float(row[3])       # hours per control
            flat = _safe_float(row[4])       # flat hours
            qa = _safe_float(row[5])         # QA hours
            bdays = _safe_int(row[6])        # business days

            # Workshop linking: match column H against workshop name map (case-insensitive)
            ws_ref = str(row[7]).strip().lower() if len(row) > 7 and row[7] else None
            ws_template_id = None
            if ws_ref and ws_ref in wt_name_map:
                ws_template_id = wt_name_map[ws_ref].id

            dt = DeliverableTemplate(
                project_template_id=pt.id,
                name=name,
                deliverable_type=dtype,
                # Only set control family fields for control_family type
                control_family_code=cf_code if dtype == "control_family" else None,
                default_hours_per_control=hpc if dtype == "control_family" else None,
                default_flat_hours=flat,
                default_qa_hours=qa,
                default_business_days=bdays,
                workshop_template_id=ws_template_id,
                sort_order=idx,
            )
            db.add(dt)

    db.commit()
    return _load_template(db, pt.id)


# ── Project Templates ──────────────────────────────────────────────────────────

@router.get("", response_model=list[ProjectTemplateOut])
def list_templates(db: DB, _: Auth):
    """GET /templates — List all project templates.

    Returns:
        List of ProjectTemplateOut objects ordered by name, with deliverable
        and workshop templates eagerly loaded.
    """
    return (
        db.query(ProjectTemplate)
        .options(
            selectinload(ProjectTemplate.deliverable_templates),
            selectinload(ProjectTemplate.workshop_templates),
        )
        .order_by(ProjectTemplate.name)
        .all()
    )


@router.post("", response_model=ProjectTemplateOut, status_code=201)
def create_template(body: ProjectTemplateCreate, db: DB, _: Auth):
    """POST /templates — Create a new empty project template.

    Request Body:
        ProjectTemplateCreate schema with name and optional description.

    Returns:
        The newly created ProjectTemplateOut (HTTP 201).
    """
    t = ProjectTemplate(**body.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@router.patch("/{template_id}", response_model=ProjectTemplateOut)
def update_template(template_id: int, body: ProjectTemplateUpdate, db: DB, _: Auth):
    """PATCH /templates/{template_id} — Update a project template's name or description.

    Path Parameters:
        template_id (int): The template to update.

    Request Body:
        ProjectTemplateUpdate schema — partial update.

    Returns:
        The updated ProjectTemplateOut, or HTTP 404 if not found.
    """
    t = _load_template(db, template_id)
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(t, field, value)
    db.commit()
    db.refresh(t)
    return t


@router.delete("/{template_id}", status_code=204)
def delete_template(template_id: int, db: DB, _: Auth):
    """DELETE /templates/{template_id} — Delete a project template.

    Path Parameters:
        template_id (int): The template to delete.

    Returns:
        HTTP 204 No Content, or HTTP 404 if not found.

    Note:
        Cascade deletes remove all child deliverable and workshop templates.
    """
    t = db.get(ProjectTemplate, template_id)
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(t)
    db.commit()


# ── Deliverable Templates ──────────────────────────────────────────────────────

@router.post("/{template_id}/deliverables", response_model=DeliverableTemplateOut, status_code=201)
def add_deliverable_template(template_id: int, body: DeliverableTemplateCreate, db: DB, _: Auth):
    """POST /templates/{template_id}/deliverables — Add a deliverable template to a project template.

    Path Parameters:
        template_id (int): The parent project template.

    Request Body:
        DeliverableTemplateCreate schema with name, type, hours, etc.

    Returns:
        The newly created DeliverableTemplateOut (HTTP 201), or HTTP 404 if template not found.
    """
    if not db.get(ProjectTemplate, template_id):
        raise HTTPException(status_code=404, detail="Template not found")
    dt = DeliverableTemplate(project_template_id=template_id, **body.model_dump())
    db.add(dt)
    db.commit()
    db.refresh(dt)
    return dt


@router.patch("/deliverables/{dt_id}", response_model=DeliverableTemplateOut)
def update_deliverable_template(dt_id: int, body: DeliverableTemplateUpdate, db: DB, _: Auth):
    """PATCH /templates/deliverables/{dt_id} — Update a deliverable template.

    Path Parameters:
        dt_id (int): The deliverable template ID to update.

    Request Body:
        DeliverableTemplateUpdate schema — partial update.

    Returns:
        The updated DeliverableTemplateOut, or HTTP 404 if not found.
    """
    dt = db.get(DeliverableTemplate, dt_id)
    if not dt:
        raise HTTPException(status_code=404, detail="Deliverable template not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(dt, field, value)
    db.commit()
    db.refresh(dt)
    return dt


@router.delete("/deliverables/{dt_id}", status_code=204)
def delete_deliverable_template(dt_id: int, db: DB, _: Auth):
    """DELETE /templates/deliverables/{dt_id} — Delete a deliverable template.

    Path Parameters:
        dt_id (int): The deliverable template ID to delete.

    Returns:
        HTTP 204 No Content, or HTTP 404 if not found.
    """
    dt = db.get(DeliverableTemplate, dt_id)
    if not dt:
        raise HTTPException(status_code=404, detail="Deliverable template not found")
    db.delete(dt)
    db.commit()


# ── Workshop Templates ─────────────────────────────────────────────────────────

@router.post("/{template_id}/workshops", response_model=WorkshopTemplateOut, status_code=201)
def add_workshop_template(template_id: int, body: WorkshopTemplateCreate, db: DB, _: Auth):
    """POST /templates/{template_id}/workshops — Add a workshop template to a project template.

    Path Parameters:
        template_id (int): The parent project template.

    Request Body:
        WorkshopTemplateCreate schema with name and optional duration.

    Returns:
        The newly created WorkshopTemplateOut (HTTP 201), or HTTP 404 if template not found.
    """
    if not db.get(ProjectTemplate, template_id):
        raise HTTPException(status_code=404, detail="Template not found")
    wt = WorkshopTemplate(project_template_id=template_id, **body.model_dump())
    db.add(wt)
    db.commit()
    db.refresh(wt)
    return wt


@router.patch("/workshops/{wt_id}", response_model=WorkshopTemplateOut)
def update_workshop_template(wt_id: int, body: WorkshopTemplateUpdate, db: DB, _: Auth):
    """PATCH /templates/workshops/{wt_id} — Update a workshop template.

    Path Parameters:
        wt_id (int): The workshop template ID to update.

    Request Body:
        WorkshopTemplateUpdate schema — partial update.

    Returns:
        The updated WorkshopTemplateOut, or HTTP 404 if not found.
    """
    wt = db.get(WorkshopTemplate, wt_id)
    if not wt:
        raise HTTPException(status_code=404, detail="Workshop template not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(wt, field, value)
    db.commit()
    db.refresh(wt)
    return wt


@router.delete("/workshops/{wt_id}", status_code=204)
def delete_workshop_template(wt_id: int, db: DB, _: Auth):
    """DELETE /templates/workshops/{wt_id} — Delete a workshop template.

    Path Parameters:
        wt_id (int): The workshop template ID to delete.

    Returns:
        HTTP 204 No Content, or HTTP 404 if not found.
    """
    wt = db.get(WorkshopTemplate, wt_id)
    if not wt:
        raise HTTPException(status_code=404, detail="Workshop template not found")
    db.delete(wt)
    db.commit()


# ── Apply Template to Project ─────────────────────────────────────────────────

@router.post("/apply/{project_id}/{template_id}")
def apply_template(project_id: int, template_id: int, db: DB, _: Auth):
    """POST /templates/apply/{project_id}/{template_id} — Apply a template to an existing project.

    Clone all DeliverableTemplates and WorkshopTemplates from the given template
    onto the specified project as real Deliverable and Workshop records.

    Path Parameters:
        project_id (int): The project to apply the template to.
        template_id (int): The template to clone from.

    Returns:
        {"deliverables_created": int, "workshops_created": int}

    Business Logic:
        1. Validate both the project and template exist.
        2. If the project has an impact_level_id, load all ControlFamily records
           for that impact level into a code -> ControlFamily map. This is used
           to resolve control_family_code from the template into actual control
           family IDs and control counts.
        3. Create Workshop records first (sorted by sort_order), building a
           workshop_template_id -> workshop_id map for deliverable linking.
        4. Create Deliverable records (sorted by sort_order):
           - For control_family type: resolve control_family_id and control_count
             from the project's impact level.
           - Link to the workshop created in step 3 if the template specifies one.
           - Auto-create standard phases (workshop, draft, qa, delivery) for each
             deliverable. Workshop phase is skipped for non-workshop types.
        5. Commit all records and return creation counts.
    """
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    template = _load_template(db, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Build a control family lookup: code -> ControlFamily for the project's impact level
    cf_map: dict[str, ControlFamily] = {}
    if project.impact_level_id:
        cfs = db.query(ControlFamily).filter(
            ControlFamily.impact_level_id == project.impact_level_id
        ).all()
        cf_map = {cf.code: cf for cf in cfs}

    # Step 1: Create workshops first, building a template_id -> real_id map
    wt_id_map: dict[int, int] = {}  # workshop_template.id -> workshop.id
    workshops_created = 0
    for wt in sorted(template.workshop_templates, key=lambda x: x.sort_order):
        w = Workshop(project_id=project_id, name=wt.name, status="scheduled",
                     duration_hours=wt.duration_hours)
        db.add(w)
        db.flush()  # Get w.id for deliverable linking
        wt_id_map[wt.id] = w.id
        workshops_created += 1

    # Step 2: Create deliverables, linking to workshops via the map
    created = 0
    for dt in sorted(template.deliverable_templates, key=lambda x: x.sort_order):
        # Resolve control family from the template's code to the project's impact level
        cf = cf_map.get(dt.control_family_code or "")
        # Resolve workshop link from template workshop ID to real workshop ID
        workshop_id = wt_id_map.get(dt.workshop_template_id) if dt.workshop_template_id else None
        d = Deliverable(
            project_id=project_id,
            name=dt.name,
            deliverable_type=dt.deliverable_type,
            control_family_id=cf.id if cf else None,
            control_count=cf.control_count if cf else None,
            hours_per_control=dt.default_hours_per_control,
            flat_hours=dt.default_flat_hours,
            qa_hours=dt.default_qa_hours,
            business_days=dt.default_business_days,
            workshop_id=workshop_id,
            sort_order=dt.sort_order,
        )
        db.add(d)
        db.flush()  # Get d.id for phase creation
        # Create standard phases: skip "workshop" phase for non-workshop deliverables
        phase_types = _PHASE_TYPES if dt.deliverable_type == "workshop" else _PHASE_TYPES[1:]
        for i, pt_type in enumerate(phase_types):
            db.add(DeliverablePhase(deliverable_id=d.id, phase_type=pt_type, sort_order=i))
        created += 1

    db.commit()
    return {"deliverables_created": created, "workshops_created": workshops_created}
